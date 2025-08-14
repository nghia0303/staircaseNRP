import os
import signal
import sys
import datetime
import gc

import psutil
from openpyxl import Workbook
from openpyxl.styles import Alignment

import time
from datetime import datetime
from typing import Any

from pysat.solvers import Solver

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../../')))

from src.encoding.nurse_roostering_encoding import NurseRosteringEncoding, NurseRosteringConfig
from src.include.addline import write_full
from src.include.common import myrange_inclusive, cl, AuxVariable, AddClause

solver = "kissat"  # kissat, cadical, glucose, glucose-syrup

solve_mode = "local_solver" # pysat or local_solver (kissat, cadical, glucose, glucose-syrup)

KISSAT_PATH = "/home/nghia/Desktop/Crew/SAT/kissat/build/kissat"

CADICAL_PATH = "/home/nghia/Desktop/Crew/SAT/cadical/build/cadical"

GLUCOSE_PATH = "/home/nghia/Desktop/Crew/SAT/glucose/simp/glucose"

GLUCOSE_SYRUP_PATH = "/home/nghia/Desktop/Crew/SAT/glucose/parallel/glucose-syrup"

TIMEOUT = 600 # seconds

def handler(signum, frame):
	raise TimeoutError("Function execution time exceeded the limit")


def run(command: str) -> int:
	print(command)
	return os.system(command)


def eval_window(x: str, value: str, window_size: int, floor, cap):
	if len(x) < window_size:
		return True
	now = 0
	if floor is None:
		floor = 0
	if cap is None:
		cap = window_size
	for i in range(0, len(x)):
		if i >= window_size:
			if not (floor <= now <= cap):
				# print(floor, now, cap)
				# print(f"failed at {i}")
				return False
			# print(f"{x[i - window_size]} {value}")
			if value == "E_N":
				now -= 1 if x[i - window_size] == "E" else 0
				now -= 1 if x[i - window_size] == "N" else 0
			else:
				now -= 1 if x[i - window_size] == value else 0

			# print(f"now: {now}")
		if value == "E_N":
			now += 1 if x[i - window_size] == "E" else 0
			now += 1 if x[i - window_size] == "N" else 0
		else:
			now += 1 if x[i] == value else 0
	if not (floor <= now <= cap):
		print(f"failed at end")
		return False
	return True


def eval_window_lower_bound(x: str, value: str, window_size: int, floor: int):
	return eval_window(x, value, window_size, floor, None)


def eval_window_upper_bound(x: str, value: str, window_size: int, cap: int):
	return eval_window(x, value, window_size, None, cap)

def solve(start_time, cannon_name: str, aux, add_clause: AddClause) -> list[str | None]:

	if not os.path.exists("tmp"):
		os.makedirs("tmp")
	cnf_file = f"tmp/{cannon_name}.cnf"
	start_write_time = time.perf_counter()
	write_full(aux.get_total_added_var(), add_clause.get_clause(), cnf_file)
	os.system(f"head -n1 {cnf_file}")
	print("Write time: {:.2f} (ms)".format((time.perf_counter() - start_write_time) * 1000))
	print("Encoding time (including writing to file): {:.2f} (ms)".format((time.perf_counter() - start_time) * 1000))
	start_solve_time = time.perf_counter()

	if not os.path.exists("tmp/solver_output"):
		os.makedirs("tmp/solver_output")
	solver_output = f"tmp/solver_output/output_{cannon_name}.txt"
	# ret = run(f"kissat -q {cnf_file} > {solver_output}")
	if solver == "kissat":
		ret = run(f"{KISSAT_PATH} -q {cnf_file} > {solver_output}")
		print(f"Solve time: {(time.perf_counter() - start_solve_time) * 1000:.2f} (ms)")
		return [
			ret, solver_output
		]
	elif solver == "cadical":
		ret = run(f"{CADICAL_PATH} -q {cnf_file} > {solver_output}")
		print(f"Solve time: {(time.perf_counter() - start_solve_time) * 1000:.2f} (ms)")
		return [
			ret, solver_output
		]
	elif solver == "glucose":
		ret = run(f"{GLUCOSE_PATH} -model -verb=0 {cnf_file} | grep -E '^(s|v) ' >> {solver_output}")
		print(f"Solve time: {(time.perf_counter() - start_solve_time) * 1000:.2f} (ms)")
		return [
			ret, solver_output
		]
	elif solver == "glucose_syrup":
		ret = run(f"{GLUCOSE_SYRUP_PATH} -model -verb=0 {cnf_file} | grep -E '^(s|v) ' >> {solver_output}")
		print(f"Solve time: {(time.perf_counter() - start_solve_time) * 1000:.2f} (ms)")
		return [
			ret, solver_output
		]
	elif "pysat_" in solver:
		solver_name = solver.split('_')[1]

		pysat_solver = Solver(name=solver_name)
		pysat_solver.solve()
	else:
		raise RuntimeError(f"unknown solver {solver}")
	print(f"Solve time: {(time.perf_counter() - start_solve_time) * 1000:.2f} (ms)")
	return [None, None]

def solve_with_pysat(start_time, cannon_name: str, aux, add_clause: AddClause) -> list[str | None]:
	# start_solve_time = time.perf_counter()
	if not os.path.exists("tmp/solver_output"):
		os.makedirs("tmp/solver_output")
	solver_output = f"tmp/solver_output/output_{cannon_name}.txt"
	cnf = add_clause.formula

	pysat_solver = Solver(
		name='g421',
		bootstrap_with=cnf
	)

	print("Total variables:", aux.get_total_added_var())
	print("Total clauses:", add_clause.get_added_clause())

	result = pysat_solver.solve()

	print(f"Result: {result}")

	model = pysat_solver.get_model()

	# print(model)

	# print(f"Solve time: {(time.perf_counter() - start_solve_time) * 1000:.2f} (ms)")

	if result:
		return [2560, solver_output, model]
	else:
		return [5120, None]

	return [None, None]

def run_nurse_rostering(name: str, nurse: int, day: int, time_limit: int):
    signal.signal(signal.SIGALRM, handler)
    signal.alarm(time_limit)
    cannon_name = f"nurse_rostering_{name}_{nurse}_{day}"

    print(f"{name} {nurse} {day}")
    aux = AuxVariable(1)
    clause = []
    add_clause = AddClause(clause)
    total_variable = -1
    total_clause = -1
    try:
        start_time = time.perf_counter()
        nr_config = NurseRosteringConfig(nurse, day, aux, add_clause, name)
        nr = NurseRosteringEncoding(nr_config)
        nr.encode()
        encoding_time = (time.perf_counter() - start_time) * 1000
        print("Encoding time: {:.2f} (ms)".format(encoding_time))

        total_variable = aux.get_total_added_var()
        total_clause = add_clause.get_added_clause()
        solver_return = ''
        model = None
        start_solving_time = time.perf_counter()
        if solve_mode == "local_solver":
            ret, solver_output = solve(start_time, cannon_name, aux, add_clause)
        else:
            ret, solver_output, model = solve_with_pysat(start_time, cannon_name, aux, add_clause)
        solving_time = (time.perf_counter() - start_solving_time) * 1000
        print("Solving time: {:.2f} (ms)".format(solving_time))
        end_time = time.perf_counter()
        elapsed_time_ms = (end_time - start_time) * 1000
        print(f"Total time: {elapsed_time_ms:.2f} (ms)")
        if solver_output is not None and model is not None:
            with open(solver_output, "w") as f:
                f.write(' '.join(map(str, model)) + '\n')
        del nr
        del clause
        gc.collect()

        ok_time = True
        if ret == 2560:  # SAT
            solver_return = 'SAT'
            print("solns: 1")
            test_result(solver_output, nurse, day)
        elif ret == 5120:  # UNSAT
            print("UNSAT")
            print("solns: 0")
            solver_return = 'UNSAT'
        else:
            if ret == 0:  # timeout
                ok_time = False

        print(f"took {elapsed_time_ms:.2f} (ms)")
        if ok_time:
            return encoding_time, solving_time, elapsed_time_ms, solver_return, total_variable, total_clause
        else:
            return None, 'timeout', total_variable, total_clause
    except TimeoutError as te:
        print(te)
        return None, None, None, 'timeout', total_variable, total_clause
    except RecursionError as re:
        print(re)
        return None, None, None, 'timeout', total_variable, total_clause
    except OSError as e:
        print(e)
        return None, None, None, 'timeout', total_variable, total_clause
    finally:
        signal.alarm(0)


def get_all_number_in_file(file: str) -> list[list[int]]:
	args = []
	file = open(f'{file}', 'r')
	for line in file:
		arg = [int(num) for num in line.split()]
		args.append(arg)
	file.close()
	return args


def number_to_column_letter(n: int) -> str:
	result = ""
	while n > 0:
		n, remainder = divmod(n - 1, 26)
		result = chr(65 + remainder) + result

	return result


def pos_2d_to_pos_excel(x: int, y: int) -> str:
	ret = number_to_column_letter(y) + str(x)
	# print(f"({x}, {y}) -> {ret}")
	return ret


def write_to_cell(cell, value: Any):
	cell.value = value
	cell.alignment = Alignment(horizontal="center")


class DataToXlsx:
	total_variable = 'total_variable'
	clause = 'clause'
	time = 'time (ms)'
	sat_status = 'sat/unsat/timeout'

	def __init__(self, excel_file_name: str, name_list: list[str]):
		if not os.path.exists("out/excel"):
			os.makedirs("out/excel")
		self.excel_file_name = "out/excel/" + excel_file_name
		if os.path.exists(self.excel_file_name):
			print(f"file {self.excel_file_name} exist! aborting...")
			sys.exit(2)
		self.book = Workbook()
		for sheet_name in self.book.sheetnames:
			self.book.remove(self.book[sheet_name])
		self.book.create_sheet('Results')
		self.sheet = self.book['Results']
		self.name_to_column: dict[str, int] = {
			'nurse': 1,
			'day': 2,
		}
		self.row_dict: dict[str, int] = {
			'nurse': 3,
			'day': 3,
		}
		self.offset_dict: dict[str, int] = {
			'nurse': 0,
			'day': 0,
			DataToXlsx.total_variable: 0,
			DataToXlsx.clause: 1,
			DataToXlsx.time: 2,
			DataToXlsx.sat_status: 3,
		}
		write_to_cell(self.sheet.cell(2, 1), 'nurse')
		write_to_cell(self.sheet.cell(2, 2), 'day')

		for i in range(len(name_list)):
			y: int = len(self.row_dict) + 1 + i * 4
			self.name_to_column[name_list[i]] = y
			self.sheet.merge_cells(f"{pos_2d_to_pos_excel(1, y)}:{pos_2d_to_pos_excel(1, y + 3)}")
			write_to_cell(self.sheet.cell(1, y), name_list[i])
			for name_offset, offset in self.offset_dict.items():
				write_to_cell(self.sheet.cell(2, y + offset), name_offset)
			self.row_dict[name_list[i]] = 3
		self.book.save(self.excel_file_name)

	def get_column(self, name: str) -> int:
		if not (name in self.name_to_column):
			raise RuntimeError(f"{name} not in name_to_column")
		return self.name_to_column.get(name)

	def get_row(self, name: str) -> int:
		ret: int = self.row_dict.get(name)
		self.row_dict[name] += 1
		return ret

	def write_more(self, name: str, result: dict[str, str | int | float]) -> None:
		row: int = self.get_row(name)
		col: int = self.get_column(name)
		for name_offset, value in result.items():
			offset: int = self.offset_dict[name_offset]
			write_to_cell(self.sheet.cell(row, col + offset), value)
		self.book.save(self.excel_file_name)


def parse_result(filename: str, nurse: int, day: int):
	with (open(filename, 'r') as file):
		list_int: list[int] = []
		for line in file.readlines():
			for seq in line.split():
				try:
					list_int.append(int(seq))
				except ValueError:
					pass
	nurse_list = []
	list_int.reverse()
	for i in range(nurse):
		nurse_day_list = []
		for j in range(day):
			nurse_day_shift_list = []
			for k in range(5):
				nurse_day_shift_list.append(list_int.pop())
			nurse_day_list.append(nurse_day_shift_list)
		nurse_list.append(nurse_day_list)
	chosen_list = []
	num_to_shift = ['D', 'E', 'N', 'O']
	for i in range(nurse):
		chosen_nurse_list = []
		for j in range(day):
			# print(f"nurse {i} day {j}: {nurse_list[i][j]}")
			# temp = [k for k in nurse_list[i][j] if (k > 0)]
			# if len(temp) != 1:
			# 	chosen_nurse_list.append((temp[0] - 1) % 4)

			temp = []
			# print(f"nurse {i} day {j}:")
			for k_idx in range(5):
				# print(f"  {k_idx}: {nurse_list[i][j][k_idx]}")
				if nurse_list[i][j][k_idx] > 0:
					temp.append(k_idx)

			if temp[0] == 4:
				raise RuntimeError(f"There was some error when encoding e_n relationship! {temp}. Maybe more than 1 shift per day per nurse?")

			if len(temp) != 1 and not(len(temp) == 2 and temp[1] == 4):
				raise RuntimeError(f"NOT satisfied 1 shift per day per nurse! {temp}")
			chosen_nurse_list.append(num_to_shift[temp[0] % 4])


		chosen_list.append(chosen_nurse_list)
	return chosen_list


def test_result(filename: str, nurse: int, day: int):
	chosen_list = parse_result(filename, nurse, day)
	# for i in range(num_cars_):
	# 	print(chosen_class[i], chosen_class_option[i])
	for nurse_id in range(nurse):
		nurse_shifts = ''
		for shift in chosen_list[nurse_id]:
			nurse_shifts += shift
		# print(f"nurse {nurse_id}: {nurse_shifts}")
		if not eval_window_lower_bound(nurse_shifts, 'O', 7, 7 - 6):
			raise RuntimeError(f"nurse id {nurse_id} failed")
		if not eval_window_lower_bound(nurse_shifts, 'O', 14, 4):
			raise RuntimeError(f"nurse id {nurse_id} failed")
		if not eval_window(nurse_shifts, 'E', 14, 4, 8):
			raise RuntimeError(f"nurse id {nurse_id} failed")
		if not eval_window_upper_bound(nurse_shifts, 'O', 28, 28 - 16):
			raise RuntimeError(f"nurse id {nurse_id} failed")
		# if not eval_window_upper_bound(nurse_shifts, 'N', 14, 4):
		# 	raise RuntimeError(f"nurse id {nurse_id} failed")
		# if not eval_window_lower_bound(nurse_shifts, 'N', 14, 1):
		# 	raise RuntimeError(f"nurse id {nurse_id} failed")
		if not eval_window(nurse_shifts, 'N', 14, 1, 4):
			raise RuntimeError(f"nurse id {nurse_id} failed")
		if not eval_window(nurse_shifts, 'E_N', 7, 2, 4):
			raise RuntimeError(f"nurse id {nurse_id} failed")
		if not eval_window_upper_bound(nurse_shifts, 'N', 2, 1):
			raise RuntimeError(f"nurse id {nurse_id} failed")
	print("ok")


def run_multiple_nurse_rostering():
	to_test: list[str] = [
		"staircase_among"
	]
	time_now = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
	nks = get_all_number_in_file("input_nurse_rostering.txt")
	excel_file_name = f"results_nurse_rostering_{time_now}.xlsx"
	writer: DataToXlsx = DataToXlsx(excel_file_name, to_test)
	timeout = 600
	for nk in nks:
		nurse = nk[0]
		day = nk[1] * 7
		writer.write_more('nurse', {'nurse': nurse})
		writer.write_more('day', {'day': day})
		for name in to_test:
			result_dict: dict[str, str | int | float] = {
				DataToXlsx.total_variable: "",
				DataToXlsx.clause: "",
				DataToXlsx.time: "",
				DataToXlsx.sat_status: ""
			}
			encoding_time, solving_time, elapsed_time_ms, solver_return, num_var, num_clause = run_nurse_rostering(name, nurse, day, timeout)
			if elapsed_time_ms is None:
				result_dict[DataToXlsx.time] = "timeout"
			else:
				result_dict[DataToXlsx.time] = f"{elapsed_time_ms:.3f}"
			if num_var == -1:
				result_dict[DataToXlsx.total_variable] = ""
				result_dict[DataToXlsx.clause] = ""
			else:
				result_dict[DataToXlsx.total_variable] = num_var
				result_dict[DataToXlsx.clause] = num_clause
				result_dict[DataToXlsx.sat_status] = solver_return

			writer.write_more(name, result_dict)
		print()


def run_single_nurse_rostering():
	process = psutil.Process(os.getpid())
	mem_before = process.memory_info().rss / (1024 * 1024)
	print(f"Memory before: {mem_before:.2f} MB")
	number_nurses = int(sys.argv[1])  # Number of nurses
	number_days = int(sys.argv[2])  # Number of days
	encoding = sys.argv[3] if len(sys.argv) > 3 else "staircase_among"
	result_file_path = sys.argv[4] if len(sys.argv) > 4 else "results_nurse_rostering.csv"

	# number_days = number_weeks * 7
	encoding_time, solving_time, elapsed_time_ms, solver_return, num_var, num_clause = run_nurse_rostering(encoding, number_nurses, number_days, TIMEOUT)
	mem_after = process.memory_info().rss / (1024 * 1024)
	print(f"Memory after: {mem_after:.2f} MB")
	print(f"Memory used: {mem_after - mem_before:.2f} MB")

	# model,encoding,clauses,var,encoding_time,solving_time,total_time
	result_line = f"{number_nurses},{number_days},Pysat,{encoding},{num_clause},{num_var},{encoding_time:.3f},{solving_time:.3f},{elapsed_time_ms:.3f}\n"

	with open(result_file_path, 'a') as result_file:
		result_file.write(result_line)

def main():
	# Add the 'src' directory to sys.path
	run_single_nurse_rostering()
	# run_multiple_nurse_rostering()
	pass


if __name__ == '__main__':
	main()
	pass
